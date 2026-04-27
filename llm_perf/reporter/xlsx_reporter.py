"""XLSX report generator using openpyxl."""

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import BaseReporter
from .utils import flatten_submodules, format_percentage, group_by_submodule_type
from llm_perf.analyzer import UnifiedResult


class XlsxReporter(BaseReporter):
    """Generate XLSX reports with multiple sheets."""

    SHEET_OVERVIEW = "总览"
    SHEET_MEMORY_BY_TYPE = "内存分解_按类型"
    SHEET_MEMORY_BY_SUBMODEL = "内存分解_按子模型"
    SHEET_ACTIVATION_BY_PHASE = "激活内存分解_按Phase"
    SHEET_COMPUTE_BY_TYPE = "计算分解_按类型"
    SHEET_PHASE_DETAIL = "Phase详情"
    SHEET_SUBMODULE_DETAIL = "子模块详情"
    SHEET_COMM_BY_PARALLEL = "通信分解_按并行方式"
    SHEET_COMM_BY_OPERATION = "通信分解_按原语"

    HEADER_FILL = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    VALUE_FONT = Font()
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

    def report(
        self,
        result: UnifiedResult,
        **kwargs,
    ) -> bytes:
        """Generate XLSX report as bytes.

        Args:
            result: UnifiedResult from UnifiedAnalyzer
            **kwargs: Additional options

        Returns:
            XLSX file content as bytes
        """
        wb = Workbook()
        result_dict = result.to_dict()

        self._create_overview_sheet(wb, result_dict)
        self._create_memory_by_type_sheet(wb, result_dict)
        self._create_memory_by_submodel_sheet(wb, result_dict)
        self._create_activation_by_phase_sheet(wb, result_dict)
        self._create_compute_by_type_sheet(wb, result_dict)
        self._create_phase_detail_sheet(wb, result_dict)
        self._create_submodule_detail_sheet(wb, result_dict)
        self._create_comm_by_parallel_sheet(wb, result_dict)
        self._create_comm_by_operation_sheet(wb, result_dict)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def save(
        self,
        result: UnifiedResult,
        path: Union[str, Path],
        **kwargs,
    ) -> None:
        """Save XLSX report to file.

        Args:
            result: UnifiedResult
            path: Output file path
            **kwargs: Additional options
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        xlsx_bytes = self.report(result, **kwargs)
        with open(path, "wb") as f:
            f.write(xlsx_bytes)

    def _create_overview_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create overview sheet with basic info and throughput."""
        ws = wb.active
        ws.title = self.SHEET_OVERVIEW

        headers = ["指标", "值"]
        self._write_header_row(ws, headers, 1)

        overview_data = [
            ("模型名称", result_dict.get("workload_name", "-")),
            ("场景类型", result_dict.get("scenario_type", result_dict.get("workload_type", "-"))),
            ("总时间", f"{result_dict.get('total_time_sec', 0):.2f} s"),
            ("峰值内存", f"{result_dict.get('peak_memory_gb', 0):.2f} GB"),
            ("吞吐量(Tokens/sec)", f"{result_dict.get('throughput', {}).get('tokens_per_sec', 0):.0f}"),
            ("吞吐量(Samples/sec)", f"{result_dict.get('throughput', {}).get('samples_per_sec', 0):.2f}"),
            ("吞吐量(Pixels/sec)", f"{result_dict.get('throughput', {}).get('pixels_per_sec', 0):.0f}"),
            ("MFU", f"{result_dict.get('mfu', 0) * 100:.1f}%" if result_dict.get("mfu") else "-"),
            ("QPS", f"{result_dict.get('qps', 0):.0f}" if result_dict.get("qps") else "-"),
        ]

        for row_idx, (label, value) in enumerate(overview_data, start=2):
            ws.cell(row=row_idx, column=1, value=label).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=1).alignment = self.LEFT_ALIGN
            ws.cell(row=row_idx, column=2, value=value).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2).alignment = self.CENTER_ALIGN

        self._auto_adjust_column_width(ws, [15, 25])

    def _create_memory_by_type_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create memory breakdown by type sheet."""
        ws = wb.create_sheet(self.SHEET_MEMORY_BY_TYPE)

        headers = ["内存类型", "大小(GB)", "占比"]
        self._write_header_row(ws, headers, 1)

        detailed = result_dict.get("detailed_breakdown", {})
        mem_by_type = detailed.get("memory", {}).get("by_type", {})

        if not mem_by_type:
            ws.cell(row=2, column=1, value="无数据")
            return

        total = mem_by_type.get("total", 0)
        priority_order = ["weight", "gradient", "optimizer", "activation"]
        other_keys = [k for k in mem_by_type.keys() if k not in priority_order and k != "total"]

        row_idx = 2
        for key in priority_order:
            if key in mem_by_type:
                value = mem_by_type[key]
                pct_str = format_percentage(value, total)
                ws.cell(row=row_idx, column=1, value=key).font = self.VALUE_FONT
                ws.cell(row=row_idx, column=2, value=f"{value:.2f}").font = self.VALUE_FONT
                ws.cell(row=row_idx, column=3, value=pct_str).font = self.VALUE_FONT
                row_idx += 1

        for key in other_keys:
            value = mem_by_type[key]
            pct_str = format_percentage(value, total)
            ws.cell(row=row_idx, column=1, value=key).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{value:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=pct_str).font = self.VALUE_FONT
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="总计").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=f"{total:.2f}").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="100.0%").font = Font(bold=True)

        self._auto_adjust_column_width(ws, [15, 12, 10])

    def _create_memory_by_submodel_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create memory breakdown by submodel sheet."""
        ws = wb.create_sheet(self.SHEET_MEMORY_BY_SUBMODEL)

        headers = ["子模型", "权重(GB)", "梯度(GB)", "优化器(GB)", "激活(GB)", "总内存(GB)"]
        self._write_header_row(ws, headers, 1)

        detailed = result_dict.get("detailed_breakdown", {})
        mem_by_submodel = detailed.get("memory", {}).get("by_submodel", {})

        if not mem_by_submodel:
            ws.cell(row=2, column=1, value="无数据")
            return

        row_idx = 2
        for name, mems in mem_by_submodel.items():
            weight = mems.get("weight_gb", 0)
            gradient = mems.get("gradient_gb", 0)
            optimizer = mems.get("optimizer_gb", 0)
            activation = mems.get("activations_gb", 0)
            total = weight + gradient + optimizer + activation

            ws.cell(row=row_idx, column=1, value=name).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{weight:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=f"{gradient:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=4, value=f"{optimizer:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=5, value=f"{activation:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=6, value=f"{total:.2f}").font = self.VALUE_FONT
            row_idx += 1

        self._auto_adjust_column_width(ws, [20, 12, 12, 12, 12, 12])

    def _create_activation_by_phase_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create activation memory breakdown by phase sheet."""
        ws = wb.create_sheet(self.SHEET_ACTIVATION_BY_PHASE)

        headers = ["Phase", "激活内存(GB)", "占比"]
        self._write_header_row(ws, headers, 1)

        detailed = result_dict.get("detailed_breakdown", {})
        by_phase_activation = detailed.get("memory", {}).get("by_phase_activation", {})

        if not by_phase_activation:
            ws.cell(row=2, column=1, value="无数据")
            self._auto_adjust_column_width(ws, [15, 15, 10])
            return

        total_activation = sum(
            data.get("activation_gb", 0) for data in by_phase_activation.values()
        )

        row_idx = 2
        for phase_name, data in by_phase_activation.items():
            activation_gb = data.get("activation_gb", 0)
            pct = (activation_gb / total_activation * 100) if total_activation > 0 else 0
            ws.cell(row=row_idx, column=1, value=phase_name).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{activation_gb:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%").font = self.VALUE_FONT
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="总计").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=f"{total_activation:.2f}").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="100.0%").font = Font(bold=True)

        self._auto_adjust_column_width(ws, [15, 15, 10])

    def _create_compute_by_type_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create compute breakdown by submodule type sheet."""
        ws = wb.create_sheet(self.SHEET_COMPUTE_BY_TYPE)

        headers = ["子模块类型", "计算时间(ms)", "计算占比", "FLOPs(GFLOPs)", "通信时间(ms)", "通信占比", "内存(GB)"]
        self._write_header_row(ws, headers, 1)

        phases = result_dict.get("phases", [])
        flat_submodules = flatten_submodules(phases)
        grouped = group_by_submodule_type(flat_submodules)

        total_compute_time = sum(
            g.get("time_sec", 0) for k, g in grouped.items() if not k.endswith("_nested")
        )
        total_comm_time = sum(
            g.get("communication_time_sec", 0) for k, g in grouped.items() if not k.endswith("_nested")
        )

        if not grouped:
            ws.cell(row=2, column=1, value="无数据")
            return

        row_idx = 2
        for sm_type, metrics in grouped.items():
            compute_time = metrics.get("time_sec", 0)
            compute_pct = format_percentage(compute_time, total_compute_time)
            flops = metrics.get("flops", 0) / 1e9
            comm_time = metrics.get("communication_time_sec", 0)
            comm_pct = format_percentage(comm_time, total_comm_time)
            memory = (
                metrics.get("weight_memory_gb", 0)
                + metrics.get("gradient_memory_gb", 0)
                + metrics.get("optimizer_memory_gb", 0)
                + metrics.get("activation_memory_gb", 0)
            )

            ws.cell(row=row_idx, column=1, value=sm_type).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{compute_time * 1000:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=compute_pct).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=4, value=f"{flops:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=5, value=f"{comm_time * 1000:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=6, value=comm_pct).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=7, value=f"{memory:.2f}").font = self.VALUE_FONT
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="总计").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=f"{total_compute_time * 1000:.2f}").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="100.0%").font = Font(bold=True)
        ws.cell(row=row_idx, column=5, value=f"{total_comm_time * 1000:.2f}").font = Font(bold=True)
        ws.cell(row=row_idx, column=6, value="100.0%").font = Font(bold=True)

        self._auto_adjust_column_width(ws, [20, 15, 10, 12, 15, 10, 12])

    def _create_phase_detail_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create phase detail sheet with complete phase data."""
        ws = wb.create_sheet(self.SHEET_PHASE_DETAIL)

        headers = [
            "Phase名称",
            "组件",
            "计算类型",
            "单次时间(ms)",
            "重复次数",
            "总时间(ms)",
            "内存(GB)",
        ]
        self._write_header_row(ws, headers, 1)

        phases = result_dict.get("phases", [])

        if not phases:
            ws.cell(row=2, column=1, value="无数据")
            return

        row_idx = 2
        for phase in phases:
            name = phase.get("name", "-")
            component = phase.get("component", "-")
            compute_type = phase.get("compute_type", "-")
            single_time = phase.get("single_time_sec", 0) * 1000
            repeat = phase.get("repeat_count", 1)
            total_time = phase.get("total_time_sec", 0) * 1000
            memory = phase.get("memory_gb", 0)

            ws.cell(row=row_idx, column=1, value=name).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=component).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=compute_type).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=4, value=f"{single_time:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=5, value=repeat).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=6, value=f"{total_time:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=7, value=f"{memory:.2f}").font = self.VALUE_FONT
            row_idx += 1

        self._auto_adjust_column_width(ws, [15, 15, 12, 15, 12, 15, 12])

    def _create_submodule_detail_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create submodule detail sheet with all flattened submodules."""
        ws = wb.create_sheet(self.SHEET_SUBMODULE_DETAIL)

        headers = [
            "Phase",
            "子模块名",
            "子模块类型",
            "计算时间(ms)",
            "FLOPs",
            "权重内存(GB)",
            "激活内存(GB)",
            "通信时间(ms)",
            "通信量(GB)",
            "是否嵌套",
            "父模块类型",
        ]
        self._write_header_row(ws, headers, 1)

        phases = result_dict.get("phases", [])
        flat_submodules = flatten_submodules(phases)

        if not flat_submodules:
            ws.cell(row=2, column=1, value="无数据")
            self._auto_adjust_column_width(ws, [12] * 11)
            return

        row_idx = 2
        for sm in flat_submodules:
            is_nested = sm.get("parent_type") is not None
            ws.cell(row=row_idx, column=1, value=sm.get("phase_name", "")).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=sm.get("name", "")).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=sm.get("submodule_type", "")).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=4, value=f"{sm.get('time_sec', 0) * 1000:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=5, value=f"{sm.get('flops', 0):.2e}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=6, value=f"{sm.get('weight_memory_gb', 0):.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=7, value=f"{sm.get('activation_memory_gb', 0):.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=8, value=f"{sm.get('communication_time_sec', 0) * 1000:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=9, value=f"{sm.get('communication_gb', 0):.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=10, value="是" if is_nested else "否").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=11, value=sm.get("parent_type", "")).font = self.VALUE_FONT
            row_idx += 1

        self._auto_adjust_column_width(ws, [12, 20, 15, 12, 12, 12, 12, 12, 12, 10, 15])

    def _create_comm_by_parallel_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create communication breakdown by parallel type sheet."""
        ws = wb.create_sheet(self.SHEET_COMM_BY_PARALLEL)

        headers = ["并行方式", "通信量(GB)", "时间(ms)"]
        self._write_header_row(ws, headers, 1)

        detailed = result_dict.get("detailed_breakdown", {})
        by_parallelism = detailed.get("communication", {}).get("by_parallelism", {})

        if not by_parallelism:
            ws.cell(row=2, column=1, value="无数据")
            self._auto_adjust_column_width(ws, [15, 15, 12])
            return

        total_bytes = 0
        total_time = 0.0
        row_idx = 2
        for parallel_type, data in by_parallelism.items():
            bytes_val = data.get("total_bytes", 0)
            time_val = data.get("total_time_sec", 0.0)
            bytes_gb = bytes_val / 1e9
            time_ms = time_val * 1000
            total_bytes += bytes_val
            total_time += time_val

            ws.cell(row=row_idx, column=1, value=parallel_type).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{bytes_gb:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=f"{time_ms:.2f}").font = self.VALUE_FONT
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="总计").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=f"{total_bytes / 1e9:.2f}").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=f"{total_time * 1000:.2f}").font = Font(bold=True)

        self._auto_adjust_column_width(ws, [15, 15, 12])

    def _create_comm_by_operation_sheet(self, wb: Workbook, result_dict: Dict[str, Any]) -> None:
        """Create communication breakdown by operation type sheet."""
        ws = wb.create_sheet(self.SHEET_COMM_BY_OPERATION)

        headers = ["通信原语", "通信量(GB)", "时间(ms)", "并行方式"]
        self._write_header_row(ws, headers, 1)

        detailed = result_dict.get("detailed_breakdown", {})
        by_operation = detailed.get("communication", {}).get("by_operation", {})

        if not by_operation:
            ws.cell(row=2, column=1, value="无数据")
            self._auto_adjust_column_width(ws, [15, 15, 12, 15])
            return

        row_idx = 2
        for op_name, data in by_operation.items():
            bytes_val = data.get("total_bytes", 0)
            time_val = data.get("total_time_sec", 0.0)
            bytes_gb = bytes_val / 1e9
            time_ms = time_val * 1000
            by_ptype = data.get("by_ptype", {})
            ptypes = ", ".join(by_ptype.keys()) if by_ptype else "-"

            ws.cell(row=row_idx, column=1, value=op_name).font = self.VALUE_FONT
            ws.cell(row=row_idx, column=2, value=f"{bytes_gb:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=3, value=f"{time_ms:.2f}").font = self.VALUE_FONT
            ws.cell(row=row_idx, column=4, value=ptypes).font = self.VALUE_FONT
            row_idx += 1

        self._auto_adjust_column_width(ws, [15, 15, 12, 15])

    def _write_header_row(self, ws: Any, headers: List[str], row: int) -> None:
        """Write header row with styling."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN

    def _auto_adjust_column_width(self, ws: Any, widths: List[int]) -> None:
        """Set column widths."""
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width