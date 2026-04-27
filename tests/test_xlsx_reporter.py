"""Tests for XlsxReporter."""

import pytest
import sys
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook

from llm_perf.analyzer import UnifiedAnalyzer
from llm_perf.hardware.cluster import Cluster
from llm_perf.hardware.device import Device
from llm_perf.hardware.topology import NetworkTopology
from llm_perf.modeling import create_model_from_config
from llm_perf.reporter import XlsxReporter
from llm_perf.strategy.base import StrategyConfig


class TestXlsxReporter:
    """Test XlsxReporter functionality."""

    def _create_test_result(self):
        """Create a test UnifiedResult."""
        device = Device.from_preset("H100-SXM-80GB")
        topology = NetworkTopology(
            name="default",
            intra_node_bandwidth_gbps=200.0,
            intra_node_latency_us=1.0,
            inter_node_bandwidth_gbps=25.0,
            inter_node_latency_us=10.0,
        )
        cluster = Cluster.create_homogeneous(device.config, 8, topology)
        strategy = StrategyConfig(tp_degree=8)
        model = create_model_from_config({"type": "llama-7b"})
        analyzer = UnifiedAnalyzer(model, device, cluster, strategy)
        return analyzer.analyze("training", batch_size=32)

    def test_xlsx_reporter_import(self):
        """Test XlsxReporter can be imported."""
        from llm_perf.reporter import XlsxReporter
        assert XlsxReporter is not None

    def test_xlsx_reporter_report_returns_bytes(self):
        """Test report() returns bytes."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

    def test_xlsx_reporter_has_all_sheets(self):
        """Test XLSX has all expected sheets."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        expected_sheets = [
            "总览",
            "内存分解_按类型",
            "内存分解_按子模型",
            "激活内存分解_按Phase",
            "计算分解_按类型",
            "Phase详情",
            "子模块详情",
            "通信分解_按并行方式",
            "通信分解_按原语",
        ]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

    def test_xlsx_overview_sheet_content(self):
        """Test overview sheet has correct content."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["总览"]

        assert ws.cell(row=1, column=1).value == "指标"
        assert ws.cell(row=1, column=2).value == "值"

        row2_label = ws.cell(row=2, column=1).value
        assert row2_label is not None

        result_dict = result.to_dict()
        assert ws.cell(row=4, column=1).value == "总时间"
        assert ws.cell(row=5, column=1).value == "峰值内存"

    def test_xlsx_memory_by_type_sheet_content(self):
        """Test memory by type sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["内存分解_按类型"]

        assert ws.cell(row=1, column=1).value == "内存类型"
        assert ws.cell(row=1, column=2).value == "大小(GB)"
        assert ws.cell(row=1, column=3).value == "占比"

        row2_type = ws.cell(row=2, column=1).value
        assert row2_type in ["weight", "gradient", "optimizer", "activation"]

    def test_xlsx_compute_by_type_sheet_content(self):
        """Test compute by type sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["计算分解_按类型"]

        expected_headers = ["子模块类型", "计算时间(ms)", "计算占比", "FLOPs(GFLOPs)", "通信时间(ms)", "通信占比", "内存(GB)"]
        for col_idx, header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_idx).value == header

    def test_xlsx_phase_detail_sheet_content(self):
        """Test phase detail sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["Phase详情"]

        expected_headers = ["Phase名称", "组件", "计算类型", "单次时间(ms)", "重复次数", "总时间(ms)", "内存(GB)"]
        for col_idx, header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_idx).value == header

        row2_phase = ws.cell(row=2, column=1).value
        assert row2_phase is not None

    def test_xlsx_submodule_detail_sheet_content(self):
        """Test submodule detail sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["子模块详情"]

        expected_headers = [
            "Phase",
            "子模块名",
            "子模块类型",
            "计算时间(ms)",
            "FLOPs",
            "权重内存(GB)",
            "激活内存(GB)",
            "通信时间(ms)",
            "通信量(GB)",
            "是否嵌套",
            "父模块类型",
        ]
        for col_idx, header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_idx).value == header

        row2_phase = ws.cell(row=2, column=1).value
        assert row2_phase is not None

    def test_xlsx_submodule_detail_sheet_has_data(self):
        """Test submodule detail sheet contains data from result."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["子模块详情"]

        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=2, column=2).value is not None
        assert ws.cell(row=2, column=3).value is not None

        time_value = ws.cell(row=2, column=4).value
        assert time_value is not None
        assert isinstance(time_value, str)
        assert float(time_value) >= 0

    def test_xlsx_comm_by_parallel_sheet_content(self):
        """Test communication by parallel type sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["通信分解_按并行方式"]

        assert ws.cell(row=1, column=1).value == "并行方式"
        assert ws.cell(row=1, column=2).value == "通信量(GB)"
        assert ws.cell(row=1, column=3).value == "时间(ms)"

    def test_xlsx_comm_by_operation_sheet_content(self):
        """Test communication by operation sheet has correct structure."""
        result = self._create_test_result()
        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["通信分解_按原语"]

        assert ws.cell(row=1, column=1).value == "通信原语"
        assert ws.cell(row=1, column=2).value == "通信量(GB)"
        assert ws.cell(row=1, column=3).value == "时间(ms)"
        assert ws.cell(row=1, column=4).value == "并行方式"

    def test_xlsx_reporter_save(self):
        """Test save() creates file."""
        import tempfile

        result = self._create_test_result()
        reporter = XlsxReporter()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        reporter.save(result, temp_path)

        assert temp_path.exists()
        assert temp_path.stat().st_size > 0

        wb = load_workbook(temp_path)
        assert "总览" in wb.sheetnames

        temp_path.unlink()

    def test_xlsx_inference_result(self):
        """Test XlsxReporter with inference result."""
        device = Device.from_preset("H100-SXM-80GB")
        topology = NetworkTopology(
            name="default",
            intra_node_bandwidth_gbps=200.0,
            intra_node_latency_us=1.0,
            inter_node_bandwidth_gbps=25.0,
            inter_node_latency_us=10.0,
        )
        cluster = Cluster.create_homogeneous(device.config, 8, topology)
        strategy = StrategyConfig(tp_degree=8)
        model = create_model_from_config({"type": "llama-7b"})
        analyzer = UnifiedAnalyzer(model, device, cluster, strategy)
        result = analyzer.analyze("autoregressive-inference", batch_size=1, prompt_len=512, generation_len=100)

        reporter = XlsxReporter()
        xlsx_bytes = reporter.report(result)

        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "总览" in wb.sheetnames
        assert "Phase详情" in wb.sheetnames


class TestReporterUtils:
    """Test reporter utility functions."""

    def test_format_bytes_gb(self):
        """Test format_bytes_gb function."""
        from llm_perf.reporter.utils import format_bytes_gb

        assert format_bytes_gb(1e9) == "1.00 GB"
        assert format_bytes_gb(0) == "0.00 GB"
        assert format_bytes_gb(2.5e9) == "2.50 GB"

    def test_format_time_ms(self):
        """Test format_time_ms function."""
        from llm_perf.reporter.utils import format_time_ms

        assert format_time_ms(1) == "1000.00 ms"
        assert format_time_ms(0) == "0.00 ms"
        assert format_time_ms(0.001) == "1.00 ms"

    def test_format_percentage(self):
        """Test format_percentage function."""
        from llm_perf.reporter.utils import format_percentage

        assert format_percentage(50, 100) == "50.0%"
        assert format_percentage(0, 100) == "0.0%"
        assert format_percentage(1, 3) == "33.3%"

    def test_flatten_submodules_empty(self):
        """Test flatten_submodules with empty phases."""
        from llm_perf.reporter.utils import flatten_submodules

        result = flatten_submodules([])
        assert result == []

    def test_group_by_submodule_type(self):
        """Test group_by_submodule_type function."""
        from llm_perf.reporter.utils import group_by_submodule_type

        flat = [
            {"submodule_type": "attention", "time_sec": 0.01, "flops": 1e9},
            {"submodule_type": "attention", "time_sec": 0.02, "flops": 2e9},
            {"submodule_type": "ffn", "time_sec": 0.03, "flops": 3e9},
        ]

        grouped = group_by_submodule_type(flat)
        assert "attention" in grouped
        assert "ffn" in grouped
        assert grouped["attention"]["time_sec"] == 0.03
        assert grouped["attention"]["flops"] == 3e9
        assert grouped["ffn"]["time_sec"] == 0.03

    def test_group_by_submodule_type_nested(self):
        """Test group_by_submodule_type with nested submodules."""
        from llm_perf.reporter.utils import group_by_submodule_type

        flat = [
            {"submodule_type": "embedding", "time_sec": 0.01, "flops": 1e9},
            {"submodule_type": "transformer_block", "time_sec": 0.05, "flops": 5e9},
            {"submodule_type": "attention", "time_sec": 0.02, "flops": 2e9, "parent_type": "transformer_block"},
            {"submodule_type": "ffn", "time_sec": 0.03, "flops": 3e9, "parent_type": "transformer_block"},
        ]

        grouped = group_by_submodule_type(flat)
        assert "embedding" in grouped
        assert "transformer_block" in grouped
        assert "attention_nested" in grouped
        assert "ffn_nested" in grouped
        
        assert grouped["embedding"]["time_sec"] == 0.01
        assert grouped["transformer_block"]["time_sec"] == 0.05
        assert grouped["attention_nested"]["time_sec"] == 0.02
        assert grouped["ffn_nested"]["time_sec"] == 0.03
        
        assert grouped["attention_nested"]["parent_type"] == "transformer_block"
        assert grouped["ffn_nested"]["parent_type"] == "transformer_block"

    def test_group_by_submodule_type_total_time_no_double_count(self):
        """Test that total time doesn't double-count nested submodules."""
        from llm_perf.reporter.utils import group_by_submodule_type

        flat = [
            {"submodule_type": "embedding", "time_sec": 0.01, "flops": 1e9},
            {"submodule_type": "transformer_block", "time_sec": 0.05, "flops": 5e9},
            {"submodule_type": "attention", "time_sec": 0.02, "flops": 2e9, "parent_type": "transformer_block"},
            {"submodule_type": "ffn", "time_sec": 0.03, "flops": 3e9, "parent_type": "transformer_block"},
        ]

        grouped = group_by_submodule_type(flat)
        
        total_compute_time = sum(
            g.get("time_sec", 0) for k, g in grouped.items() if not k.endswith("_nested")
        )
        
        assert total_compute_time == pytest.approx(0.06)
        assert total_compute_time < grouped["transformer_block"]["time_sec"] + grouped["attention_nested"]["time_sec"] + grouped["ffn_nested"]["time_sec"]


class TestWebXlsxExport:
    """Test web API XLSX export endpoint."""

    def test_export_xlsx_endpoint(self):
        """Test /api/export/xlsx endpoint."""
        from web.app import app

        client = app.test_client()

        eval_response = client.post(
            "/api/evaluate",
            json={
                "model": {"type": "llama-7b"},
                "device": "H100-SXM-80GB",
                "num_devices": 8,
                "devices_per_node": 8,
                "topology": {"type": "2-Tier Simple", "intra_node_bw_gbps": 200},
                "strategy": {"tp": 8, "pp": 1, "dp": 1},
                "workload": {"scenario": "training", "batch_size": 32},
            },
        )

        assert eval_response.status_code == 200
        eval_data = eval_response.get_json()
        assert eval_data["success"] is True

        export_response = client.post(
            "/api/export/xlsx",
            json={"result": eval_data["result"]},
        )

        assert export_response.status_code == 200
        assert export_response.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        xlsx_bytes = export_response.data
        assert len(xlsx_bytes) > 0

        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "总览" in wb.sheetnames

    def test_export_xlsx_no_result(self):
        """Test /api/export/xlsx with no result data."""
        from web.app import app

        client = app.test_client()

        response = client.post(
            "/api/export/xlsx",
            json={},
        )

        assert response.status_code == 400
        data = response.get_json()
        assert data["success"] is False

    def test_export_xlsx_inference_result(self):
        """Test /api/export/xlsx with inference result."""
        from web.app import app

        client = app.test_client()

        eval_response = client.post(
            "/api/evaluate",
            json={
                "model": {"type": "llama-7b"},
                "device": "H100-SXM-80GB",
                "num_devices": 8,
                "devices_per_node": 8,
                "topology": {"type": "2-Tier Simple", "intra_node_bw_gbps": 200},
                "strategy": {"tp": 8, "pp": 1, "dp": 1},
                "workload": {"scenario": "inference", "batch_size": 1, "input_tokens": 512, "output_tokens": 100},
            },
        )

        assert eval_response.status_code == 200
        eval_data = eval_response.get_json()

        export_response = client.post(
            "/api/export/xlsx",
            json={"result": eval_data["result"]},
        )

        assert export_response.status_code == 200

        wb = load_workbook(BytesIO(export_response.data))
        ws = wb["总览"]
        assert ws.cell(row=1, column=1).value == "指标"